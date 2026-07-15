"""
Genera un archivo Excel de ejemplo con resultados de laboratorio enfocados en ACR y TFG.

Ejecutar con: python -m backend.seed.create_example_excel
"""

import sys
from pathlib import Path
from datetime import date

# Agregar el directorio raíz al path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

try:
    import openpyxl
except ImportError:
    print("❌ Se requiere openpyxl. Instalar con: pip install openpyxl")
    sys.exit(1)


# Columnas esperadas por el sistema
COLUMNS = [
    "Identificacion_Paciente",
    "Nombre_Paciente",
    "Apellido_Paciente",
    "Fecha_Nacimiento",
    "Sexo",
    "Telefono_Paciente",
    "Email_Paciente",
    "WhatsApp_Paciente",
    "Cedula_Medico",
    "Nombre_Medico",
    "Especialidad_Medico",
    "Codigo_Prueba",
    "Valor",
    "Fecha_Toma",
    "Fecha_Resultado",
    "Observaciones",
]


def create_example_excel():
    """Genera el archivo Excel de ejemplo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Encabezados
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)

    fecha_toma = date(2026, 7, 10)
    fecha_resultado = date(2026, 7, 10)

    # ---------------------------------------------------------------------------
    # Datos de ejemplo focalizados en ACR y TFG
    # ---------------------------------------------------------------------------
    rows = [
        # --- Paciente 1: María Elena García — Valores Normales ---
        {
            "Identificacion_Paciente": "GARM850315HDFRRL09",
            "Nombre_Paciente": "María Elena",
            "Apellido_Paciente": "García Ramírez",
            "Fecha_Nacimiento": "1985-03-15",
            "Sexo": "F",
            "Telefono_Paciente": "+52 55 1234 5678",
            "Email_Paciente": "maria.garcia@email.com",
            "WhatsApp_Paciente": "+52 55 1234 5678",
            "Cedula_Medico": "12345678",
            "Nombre_Medico": "Dr. Carlos Alberto López Hernández",
            "Especialidad_Medico": "Medicina General",
            "Codigo_Prueba": "CRTS",
            "Valor": 0.8,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "",
        },
        {
            "Identificacion_Paciente": "GARM850315HDFRRL09",
            "Nombre_Paciente": "María Elena",
            "Apellido_Paciente": "García Ramírez",
            "Fecha_Nacimiento": "1985-03-15",
            "Sexo": "F",
            "Telefono_Paciente": "+52 55 1234 5678",
            "Email_Paciente": "maria.garcia@email.com",
            "WhatsApp_Paciente": "+52 55 1234 5678",
            "Cedula_Medico": "12345678",
            "Nombre_Medico": "Dr. Carlos Alberto López Hernández",
            "Especialidad_Medico": "Medicina General",
            "Codigo_Prueba": "ALBOR",
            "Valor": 12.0,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "",
        },
        {
            "Identificacion_Paciente": "GARM850315HDFRRL09",
            "Nombre_Paciente": "María Elena",
            "Apellido_Paciente": "García Ramírez",
            "Fecha_Nacimiento": "1985-03-15",
            "Sexo": "F",
            "Telefono_Paciente": "+52 55 1234 5678",
            "Email_Paciente": "maria.garcia@email.com",
            "WhatsApp_Paciente": "+52 55 1234 5678",
            "Cedula_Medico": "12345678",
            "Nombre_Medico": "Dr. Carlos Alberto López Hernández",
            "Especialidad_Medico": "Medicina General",
            "Codigo_Prueba": "CRE01",
            "Valor": 110.0,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "",
        },
        {
            "Identificacion_Paciente": "GARM850315HDFRRL09",
            "Nombre_Paciente": "María Elena",
            "Apellido_Paciente": "García Ramírez",
            "Fecha_Nacimiento": "1985-03-15",
            "Sexo": "F",
            "Telefono_Paciente": "+52 55 1234 5678",
            "Email_Paciente": "maria.garcia@email.com",
            "WhatsApp_Paciente": "+52 55 1234 5678",
            "Cedula_Medico": "12345678",
            "Nombre_Medico": "Dr. Carlos Alberto López Hernández",
            "Especialidad_Medico": "Medicina General",
            "Codigo_Prueba": "ACR",
            "Valor": 10.9,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "",
        },
        # --- Paciente 2: Pedro López — Valores Alterados ---
        {
            "Identificacion_Paciente": "LOPC900728HDFPRS05",
            "Nombre_Paciente": "Pedro",
            "Apellido_Paciente": "López Cruz",
            "Fecha_Nacimiento": "28/07/1990",
            "Sexo": "M",
            "Telefono_Paciente": "+52 55 2345 6789",
            "Email_Paciente": "pedro.lopez@email.com",
            "WhatsApp_Paciente": "+52 55 2345 6789",
            "Cedula_Medico": "23456789",
            "Nombre_Medico": "Dra. María Fernanda Gutiérrez Sánchez",
            "Especialidad_Medico": "Medicina Interna",
            "Codigo_Prueba": "CRTS",
            "Valor": 1.8,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "Paciente reporta ayuno de 8 horas",
        },
        {
            "Identificacion_Paciente": "LOPC900728HDFPRS05",
            "Nombre_Paciente": "Pedro",
            "Apellido_Paciente": "López Cruz",
            "Fecha_Nacimiento": "28/07/1990",
            "Sexo": "M",
            "Telefono_Paciente": "+52 55 2345 6789",
            "Email_Paciente": "pedro.lopez@email.com",
            "WhatsApp_Paciente": "+52 55 2345 6789",
            "Cedula_Medico": "23456789",
            "Nombre_Medico": "Dra. María Fernanda Gutiérrez Sánchez",
            "Especialidad_Medico": "Medicina Interna",
            "Codigo_Prueba": "ALBOR",
            "Valor": 35.0,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "",
        },
        {
            "Identificacion_Paciente": "LOPC900728HDFPRS05",
            "Nombre_Paciente": "Pedro",
            "Apellido_Paciente": "López Cruz",
            "Fecha_Nacimiento": "28/07/1990",
            "Sexo": "M",
            "Telefono_Paciente": "+52 55 2345 6789",
            "Email_Paciente": "pedro.lopez@email.com",
            "WhatsApp_Paciente": "+52 55 2345 6789",
            "Cedula_Medico": "23456789",
            "Nombre_Medico": "Dra. María Fernanda Gutiérrez Sánchez",
            "Especialidad_Medico": "Medicina Interna",
            "Codigo_Prueba": "CRE01",
            "Valor": 80.0,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "",
        },
        {
            "Identificacion_Paciente": "LOPC900728HDFPRS05",
            "Nombre_Paciente": "Pedro",
            "Apellido_Paciente": "López Cruz",
            "Fecha_Nacimiento": "28/07/1990",
            "Sexo": "M",
            "Telefono_Paciente": "+52 55 2345 6789",
            "Email_Paciente": "pedro.lopez@email.com",
            "WhatsApp_Paciente": "+52 55 2345 6789",
            "Cedula_Medico": "23456789",
            "Nombre_Medico": "Dra. María Fernanda Gutiérrez Sánchez",
            "Especialidad_Medico": "Medicina Interna",
            "Codigo_Prueba": "ACR",
            "Valor": 43.7,
            "Fecha_Toma": fecha_toma.isoformat(),
            "Fecha_Resultado": fecha_resultado.isoformat(),
            "Observaciones": "",
        }
    ]

    # Escribir filas
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    # Ajustar anchos de columna
    for col_idx, col_name in enumerate(COLUMNS, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            if len(cell_val) > max_len:
                max_len = len(cell_val)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
            max_len + 2, 40
        )

    # Guardar
    output_path = Path(__file__).parent / "ejemplo_resultados_corregido.xlsx"
    wb.save(str(output_path))
    print(f"✅ Archivo Excel creado: {output_path}")
    print(f"   Total de filas: {len(rows)}")

if __name__ == "__main__":
    create_example_excel()
